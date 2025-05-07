
import openpyxl
import os

DEFAULT_SHEETS = ["material", "furniture", "decorative"]
RATE_CARD_FILE = "rate_card.xlsx"

class RateCardExcel:
    def __init__(self, filepath=RATE_CARD_FILE):
        self.filepath = filepath
        if not os.path.exists(self.filepath):
            self._create_default_file()
        self.workbook = openpyxl.load_workbook(self.filepath)

    def _create_default_file(self):
        wb = openpyxl.Workbook()
        for idx, sheet in enumerate(DEFAULT_SHEETS):
            ws = wb.create_sheet(title=sheet.capitalize(), index=idx)
            ws.append(["Name", "UOM", "Rate"])
        del wb["Sheet"]
        wb.save(self.filepath)

    def get_items(self, category):
        sheet = self.workbook[category.capitalize()]
        items = []
        for row in sheet.iter_rows(min_row=2, values_only=True):
            if row and row[0]:
                items.append({"name": row[0], "uom": row[1], "rate": row[2]})
        return items

    def add_item(self, category, name, uom, rate):
        sheet = self.workbook[category.capitalize()]
        sheet.append([name, uom, rate])
        self.workbook.save(self.filepath)

    def update_item(self, category, index, name, uom, rate):
        sheet = self.workbook[category.capitalize()]
        row = index + 2  # account for header
        if 2 <= row <= sheet.max_row:
            sheet[f"A{row}"] = name
            sheet[f"B{row}"] = uom
            sheet[f"C{row}"] = rate
            self.workbook.save(self.filepath)

    def delete_item(self, category, index):
        sheet = self.workbook[category.capitalize()]
        row = index + 2
        if 2 <= row <= sheet.max_row:
            sheet.delete_rows(row, 1)
            self.workbook.save(self.filepath)
