
from openpyxl import Workbook
from openpyxl.styles import Font

def export_quote_to_excel(filepath, project, rooms, quote_calculator):
    wb = Workbook()
    ws = wb.active
    ws.title = "Quote Summary"

    # Project info
    ws.append(["Client Name", project.get("client_name", "")])
    ws.append(["Location", project.get("location", "")])
    ws.append(["Project Type", project.get("project_type", "")])
    ws.append([])
    
    header_font = Font(bold=True)
    
    for room in rooms:
        summary = quote_calculator.generate_room_summary(room)
        ws.append([f"Room: {summary['room_name']}"])
        ws.append(["Item", "Category", "UOM", "Qty", "Unit Rate", "Add-Ons", "Total Cost"])
        for cell in ws[ws.max_row]:
            cell.font = header_font

        for item in summary["items"]:
            addons_str = ", ".join([f"{k}+{v}" for k, v in item.get("addons", {}).items()])
            ws.append([
                item["name"],
                item["category"],
                item.get("uom", ""),
                item["quantity"],
                item["unit_cost"],
                addons_str,
                item["total_cost"]
            ])
        ws.append(["", "", "", "", "", "Room Total", summary["total"]])
        ws.append([])

    grand_total = quote_calculator.calculate_project_total(rooms)
    ws.append(["", "", "", "", "", "Grand Total", grand_total])
    
    wb.save(filepath)
